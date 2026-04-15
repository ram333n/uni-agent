import json
import os
import re
import shutil
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, url_for

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-prod")

BASE_DIR = Path(__file__).parent
app.config["DATASETS_DIR"] = BASE_DIR / "datasets"
app.config["DOCS_DIR"] = BASE_DIR / "docs"

app.config["DATASETS_DIR"].mkdir(exist_ok=True)
app.config["DOCS_DIR"].mkdir(exist_ok=True)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
DATASET_NAME_RE = re.compile(r"^[a-zA-Z0-9_-]+$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def dataset_path(name: str) -> Path:
    return app.config["DATASETS_DIR"] / f"{name}.json"


def load_dataset(name: str) -> dict:
    path = dataset_path(name)
    if not path.exists():
        abort(404)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_dataset(data: dict) -> None:
    path = dataset_path(data["name"])
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def list_datasets() -> list[dict]:
    datasets = []
    for p in sorted(app.config["DATASETS_DIR"].glob("*.json")):
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            datasets.append({
                "name": data.get("name", p.stem),
                "question_count": len(data.get("questions", [])),
                "created_at": data.get("created_at", ""),
            })
        except (json.JSONDecodeError, OSError):
            continue
    return datasets


# ---------------------------------------------------------------------------
# Main page
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    datasets = list_datasets()
    return render_template("index.html", datasets=datasets)


@app.route("/datasets", methods=["POST"])
def create_dataset():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Dataset name is required.", "error")
        return redirect(url_for("index"))
    if not DATASET_NAME_RE.match(name):
        flash("Dataset name may only contain letters, numbers, hyphens, and underscores.", "error")
        return redirect(url_for("index"))
    if dataset_path(name).exists():
        flash(f"Dataset '{name}' already exists.", "error")
        return redirect(url_for("index"))
    data = {
        "name": name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "questions": [],
    }
    try:
        save_dataset(data)
        flash(f"Dataset '{name}' created successfully.", "success")
    except OSError as e:
        flash(f"Failed to create dataset: {e}", "error")
    return redirect(url_for("dataset_page", name=name))


@app.route("/datasets/<name>/delete", methods=["POST"])
def delete_dataset(name: str):
    path = dataset_path(name)
    if not path.exists():
        flash(f"Dataset '{name}' not found.", "error")
        return redirect(url_for("index"))
    try:
        path.unlink()
        flash(f"Dataset '{name}' deleted.", "success")
    except OSError as e:
        flash(f"Failed to delete dataset: {e}", "error")
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# Dataset page
# ---------------------------------------------------------------------------

@app.route("/datasets/<name>")
def dataset_page(name: str):
    data = load_dataset(name)
    return render_template("dataset.html", dataset=data)


# ---------------------------------------------------------------------------
# Question routes
# ---------------------------------------------------------------------------

@app.route("/datasets/<name>/questions/new")
def new_question(name: str):
    load_dataset(name)  # 404 if missing
    question = {
        "id": str(uuid.uuid4()),
        "question": "",
        "documents": [],
        "labels": [],
        "comment": "",
    }
    return render_template("question_form.html", dataset_name=name, question=question, is_new=True)


@app.route("/datasets/<name>/questions", methods=["POST"])
def create_question(name: str):
    data = load_dataset(name)
    question = _parse_question_form(request.form)
    question["id"] = str(uuid.uuid4())
    data["questions"].append(question)
    try:
        save_dataset(data)
        flash("Question added successfully.", "success")
    except OSError as e:
        flash(f"Failed to save question: {e}", "error")
    return redirect(url_for("dataset_page", name=name))


@app.route("/datasets/<name>/questions/<qid>/edit")
def edit_question(name: str, qid: str):
    data = load_dataset(name)
    question = next((q for q in data["questions"] if q["id"] == qid), None)
    if question is None:
        flash("Question not found.", "error")
        return redirect(url_for("dataset_page", name=name))
    return render_template("question_form.html", dataset_name=name, question=question, is_new=False)


@app.route("/datasets/<name>/questions/<qid>", methods=["POST"])
def update_question(name: str, qid: str):
    data = load_dataset(name)
    idx = next((i for i, q in enumerate(data["questions"]) if q["id"] == qid), None)
    if idx is None:
        flash("Question not found.", "error")
        return redirect(url_for("dataset_page", name=name))
    updated = _parse_question_form(request.form)
    updated["id"] = qid
    data["questions"][idx] = updated
    try:
        save_dataset(data)
        flash("Question updated successfully.", "success")
    except OSError as e:
        flash(f"Failed to update question: {e}", "error")
    return redirect(url_for("dataset_page", name=name))


@app.route("/datasets/<name>/questions/<qid>/delete", methods=["POST"])
def delete_question(name: str, qid: str):
    data = load_dataset(name)
    before = len(data["questions"])
    data["questions"] = [q for q in data["questions"] if q["id"] != qid]
    if len(data["questions"]) == before:
        flash("Question not found.", "error")
        return redirect(url_for("dataset_page", name=name))
    try:
        save_dataset(data)
        flash("Question deleted.", "success")
    except OSError as e:
        flash(f"Failed to delete question: {e}", "error")
    return redirect(url_for("dataset_page", name=name))


def _parse_question_form(form) -> dict:
    question_text = form.get("question", "").strip()
    documents = [d.strip() for d in form.getlist("documents") if d.strip()]
    labels = [l.strip() for l in form.getlist("labels") if l.strip()]
    comment = form.get("comment", "").strip()
    return {
        "question": question_text,
        "documents": documents,
        "labels": labels,
        "comment": comment,
    }


# ---------------------------------------------------------------------------
# File preview routes
# ---------------------------------------------------------------------------

@app.route("/preview")
def preview():
    return render_template("preview_panel.html")


@app.route("/api/files")
def api_files():
    docs_dir = app.config["DOCS_DIR"]
    files = []
    for p in sorted(docs_dir.iterdir()):
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append({"name": p.name, "ext": p.suffix.lower()})
    return jsonify(files)


@app.route("/files/<path:filename>")
def serve_file(filename: str):
    docs_dir = app.config["DOCS_DIR"]
    file_path = (docs_dir / filename).resolve()
    # Security: ensure the resolved path is inside docs_dir
    if not str(file_path).startswith(str(docs_dir.resolve())):
        abort(403)
    if not file_path.exists():
        abort(404)
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return send_file(file_path, mimetype="application/pdf")
    if ext == ".docx":
        return _render_docx(file_path)
    if ext == ".xlsx":
        return _render_xlsx(file_path)
    abort(415)


@app.route("/api/preview-file", methods=["POST"])
def preview_file_upload():
    f = request.files.get("file")
    if not f:
        abort(400)
    filename = Path(f.filename).name
    ext = Path(filename).suffix.lower()
    if ext not in {".docx", ".xlsx"}:
        abort(415)
    tmp_dir = Path(tempfile.mkdtemp())
    try:
        tmp_path = tmp_dir / filename
        f.save(tmp_path)
        if ext == ".docx":
            return _render_docx(tmp_path)
        return _render_xlsx(tmp_path)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _render_docx(path: Path):
    try:
        from docx import Document
        doc = Document(str(path))
        paragraphs = [f"<p>{p.text}</p>" for p in doc.paragraphs if p.text.strip()]
        html = "\n".join(paragraphs) or "<p><em>No text content found.</em></p>"
        return render_template("_file_content.html", content=html, filename=path.name)
    except ImportError:
        return "<p class='error'>python-docx is not installed.</p>", 500
    except Exception as e:
        return f"<p class='error'>Failed to read file: {e}</p>", 500


def _render_xlsx(path: Path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets_html = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = "".join(f"<th>{c if c is not None else ''}</th>" for c in rows[0])
            body_rows = "".join(
                "<tr>" + "".join(f"<td>{c if c is not None else ''}</td>" for c in row) + "</tr>"
                for row in rows[1:]
            )
            sheets_html.append(
                f"<h3 class='sheet-name'>{sheet}</h3>"
                f"<div class='table-wrap'><table><thead><tr>{header}</tr></thead>"
                f"<tbody>{body_rows}</tbody></table></div>"
            )
        html = "\n".join(sheets_html) or "<p><em>Workbook is empty.</em></p>"
        return render_template("_file_content.html", content=html, filename=path.name)
    except ImportError:
        return "<p class='error'>openpyxl is not installed.</p>", 500
    except Exception as e:
        return f"<p class='error'>Failed to read file: {e}</p>", 500


# ---------------------------------------------------------------------------
# 404 handler
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


if __name__ == "__main__":
    app.run(debug=True)
